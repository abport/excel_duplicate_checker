# This line imports the openpyxl library, which is used to work with Excel files
import openpyxl

# File paths for the two Excel files
# This line sets the file path for the first Excel file
file1 = "path/to/file1.xlsx"
# This line sets the file path for the second Excel file
file2 = "path/to/file2.xlsx"

# Open the first Excel file
wb1 = openpyxl.load_workbook(file1)  # This line opens the first Excel file
sheet1 = wb1.active  # This line sets the active sheet in the first Excel file

# Open the second Excel file
wb2 = openpyxl.load_workbook(file2)  # This line opens the second Excel file
sheet2 = wb2.active  # This line sets the active sheet in the second Excel file

# Iterate through the rows in the first Excel file
for row in sheet1.iter_rows():  # This line starts a loop that will go through each row in the first Excel file
    # This line creates a list with the value of each cell in the current row
    row_values = [cell.value for cell in row]
    present = True  # This line sets the present variable to True
    # This line get the first cell of the current row, and it will get the column index of that cell
    col_idx = row[0].column
    # Iterate through the columns in the second sheet
    # This line starts a loop that will go through each column in the second sheet
    for col in sheet2.iter_cols(min_col=col_idx, max_col=col_idx, values_only=True):
        for value in row_values:  # This line starts a loop that will go through each value in the row
            # Check if the cell value is present in the second Excel file
            if not value in col:  # This line checks if the current value is not in the current column
                present = False  # if value is not present, it will set present to False
                break  # and it will break the loop
        if not present:
            break
    if present:
        # This line will print that values in row is found in the second Excel file and print the entire row value
        print(
            f'Values in row {row[0].row} found in {file2} and the row value is : {row_values}')
    else:
        # This line will print that values in row is not found in the second Excel file
        print(f'Values in row {row[0].row} not found in {file2}')

# Close the Excel files
wb1.close()  # This line closes the first Excel file
wb2.close()  # This line closes the second Excel file
